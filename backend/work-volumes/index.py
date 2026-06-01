import json
import os
import io
import psycopg2

SCHEMA = os.environ.get("MAIN_DB_SCHEMA", "t_p29757712_construction_project")
CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "GET, POST, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type, X-Auth-Token",
    "Content-Type": "application/json",
}


def get_conn():
    return psycopg2.connect(os.environ["DATABASE_URL"])


def get_user(conn, token: str):
    cur = conn.cursor()
    cur.execute(
        f"""SELECT u.id, u.name, u.role FROM {SCHEMA}.sessions s
            JOIN {SCHEMA}.users u ON u.id = s.user_id
            WHERE s.token = %s AND s.expires_at > NOW()""",
        (token,),
    )
    row = cur.fetchone()
    cur.close()
    if not row:
        return None
    return {"id": row[0], "name": row[1], "role": row[2]}


def resp(status, body, extra_headers=None):
    h = {**CORS}
    if extra_headers:
        h.update(extra_headers)
    return {"statusCode": status, "headers": h, "body": json.dumps(body, ensure_ascii=False, default=str) if isinstance(body, (dict, list)) else body}


def build_excel(rows):
    """Строим XLSX вручную через openpyxl"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Объёмы работ"

    # Заголовки
    headers = ["№", "Дата", "Прораб", "Вид работ", "Ед. изм.", "Объём", "Место", "Примечание"]
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="F97316", size=10)
    thin = Side(style="thin", color="374151")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 22

    # Данные
    row_fill_even = PatternFill("solid", fgColor="111827")
    row_fill_odd = PatternFill("solid", fgColor="1A1F2E")
    data_font = Font(color="E5E7EB", size=10)

    # Нарастающий итог: группируем по виду работ
    cumulative = {}

    for i, row in enumerate(rows, 1):
        work_name = row[3]
        volume = float(row[5])
        cumulative[work_name] = cumulative.get(work_name, 0) + volume

        values = [i, str(row[1]), row[2], row[3], row[4], volume, row[6] or "", row[7] or ""]
        fill = row_fill_even if i % 2 == 0 else row_fill_odd
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=i + 1, column=col_idx, value=val)
            cell.fill = fill
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[i + 1].height = 18

    # Итоговый лист нарастающего итога
    ws2 = wb.create_sheet("Нарастающий итог")
    ws2.cell(row=1, column=1, value="Вид работ").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=1).border = border
    ws2.cell(row=1, column=2, value="Итого объём").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=2).border = border
    ws2.cell(row=1, column=3, value="Ед. изм.").font = header_font
    ws2.cell(row=1, column=3).fill = header_fill
    ws2.cell(row=1, column=3).border = border

    # Собираем уникальные единицы по виду работ
    units_map = {}
    for row in rows:
        units_map[row[3]] = row[4]

    for i, (wname, total) in enumerate(sorted(cumulative.items()), 2):
        fill = row_fill_even if i % 2 == 0 else row_fill_odd
        ws2.cell(row=i, column=1, value=wname).font = data_font
        ws2.cell(row=i, column=1).fill = fill
        ws2.cell(row=i, column=1).border = border
        ws2.cell(row=i, column=2, value=round(total, 3)).font = data_font
        ws2.cell(row=i, column=2).fill = fill
        ws2.cell(row=i, column=2).border = border
        ws2.cell(row=i, column=3, value=units_map.get(wname, "")).font = data_font
        ws2.cell(row=i, column=3).fill = fill
        ws2.cell(row=i, column=3).border = border

    # Ширина колонок
    col_widths = [5, 12, 20, 40, 10, 10, 25, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def handler(event: dict, context) -> dict:
    """Объёмы выполненных работ: добавление, список, экспорт Excel"""

    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS, "body": ""}

    headers = event.get("headers") or {}
    token = headers.get("X-Auth-Token") or headers.get("x-auth-token") or ""
    method = event.get("httpMethod", "GET")
    params = event.get("queryStringParameters") or {}
    action = params.get("action", "")

    conn = get_conn()
    user = get_user(conn, token)
    if not user:
        conn.close()
        return resp(401, {"error": "Не авторизован"})

    # POST ?action=add — добавить строку объёма
    if method == "POST" and action == "add":
        body = json.loads(event.get("body") or "{}")
        work_name = (body.get("work_name") or "").strip()
        unit = (body.get("unit") or "").strip()
        volume_raw = body.get("volume")
        location = (body.get("location") or "").strip()
        note = (body.get("note") or "").strip()
        work_date = (body.get("work_date") or "").strip()

        if not work_name or not unit or volume_raw is None:
            conn.close()
            return resp(400, {"error": "Заполните вид работ, единицу и объём"})

        volume = float(volume_raw)

        cur = conn.cursor()
        cur.execute(
            f"""INSERT INTO {SCHEMA}.work_volumes
                (user_id, user_name, work_date, work_name, unit, volume, location, note)
                VALUES (%s, %s, %s::date, %s, %s, %s, %s, %s)
                RETURNING id, work_date, created_at""",
            (user["id"], user["name"], work_date or None, work_name, unit, volume, location, note),
        )
        row = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        return resp(201, {"id": row[0], "work_date": str(row[1]), "created_at": row[2]})

    # GET ?action=list — список всех записей
    if method == "GET" and action == "list":
        cur = conn.cursor()
        cur.execute(
            f"""SELECT id, work_date, user_name, work_name, unit, volume, location, note, created_at
                FROM {SCHEMA}.work_volumes ORDER BY work_date DESC, created_at DESC"""
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
        items = [
            {"id": r[0], "work_date": str(r[1]), "user_name": r[2], "work_name": r[3],
             "unit": r[4], "volume": float(r[5]), "location": r[6], "note": r[7], "created_at": r[8]}
            for r in rows
        ]
        # Нарастающий итог по видам работ
        cumulative = {}
        for item in reversed(items):
            cumulative[item["work_name"]] = cumulative.get(item["work_name"], 0) + item["volume"]
        return resp(200, {"items": items, "cumulative": cumulative})

    # GET ?action=export — скачать Excel
    if method == "GET" and action == "export":
        import base64
        cur = conn.cursor()
        cur.execute(
            f"""SELECT id, work_date, user_name, work_name, unit, volume, location, note
                FROM {SCHEMA}.work_volumes ORDER BY work_date ASC, created_at ASC"""
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()

        xlsx_bytes = build_excel(rows)
        b64 = base64.b64encode(xlsx_bytes).decode()

        return {
            "statusCode": 200,
            "headers": {
                **CORS,
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Content-Disposition": 'attachment; filename="work_volumes.xlsx"',
            },
            "body": b64,
            "isBase64Encoded": True,
        }

    conn.close()
    return resp(400, {"error": "Неизвестный action"})
